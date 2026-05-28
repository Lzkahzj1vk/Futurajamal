from flask import Flask, render_template, request, redirect, url_for, jsonify, flash, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta
from sqlalchemy import func
import os
import csv
from io import StringIO, BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas as rl_canvas

app = Flask(__name__)

# ─── Configuration ───
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'futura-secret-key-change-in-prod')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL',
    'postgresql://postgres:root@localhost:5432/postgres'
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# ─────────────────────────────────────────
# CONTEXT PROCESSOR
# ─────────────────────────────────────────
@app.context_processor
def inject_notifications():
    if current_user.is_authenticated:
        notifs = Notification.query.filter_by(user_id=current_user.id)\
           .order_by(Notification.created_at.desc()).limit(10).all()
        notif_unread = sum(1 for n in notifs if not n.lu)
        return dict(notifs=notifs, notif_unread=notif_unread)
    return dict(notifs=[], notif_unread=0)

# ─────────────────────────────────────────
# MODELS
# ─────────────────────────────────────────

class User(UserMixin, db.Model):
    __tablename__ = 'users'

    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)
    prenom = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(50), default='consultant')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    saisies = db.relationship('Saisie', backref='user', lazy=True, cascade='all, delete-orphan')
    conges = db.relationship('Conge', backref='user', lazy=True, cascade='all, delete-orphan')

    @property
    def initials(self):
        return (self.prenom[0] + self.nom[0]).upper() if self.prenom and self.nom else '??'

    @property
    def full_name(self):
        return f"{self.prenom} {self.nom}"

class Projet(db.Model):
    __tablename__ = 'projets'

    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(200), nullable=False)
    client = db.Column(db.String(150))
    statut = db.Column(db.String(50), default='actif')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    saisies = db.relationship('Saisie', backref='projet', lazy=True)

    @property
    def display_name(self):
        if self.client:
            return f"{self.client} — {self.nom}"
        return self.nom

class Saisie(db.Model):
    __tablename__ = 'saisies'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    projet_id = db.Column(db.Integer, db.ForeignKey('projets.id'))
    date = db.Column(db.Date, nullable=False, default=date.today)
    type_activite = db.Column(db.String(100), nullable=False)
    heures = db.Column(db.Numeric(4, 1), nullable=False)
    commentaire = db.Column(db.Text)
    statut = db.Column(db.String(50), default='en_attente')
    conge_date_debut = db.Column(db.Date, nullable=True)
    conge_date_fin = db.Column(db.Date, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    BADGE_MAP = {
        # Plans & Schémas
        'Liste des plans': 'badge-blue',
        'Schémas des tableaux': 'badge-blue',
        'Plans des locaux techniques': 'badge-blue',
        'Plans CDC': 'badge-blue',
        'Plans des locaux types': 'badge-blue',
        # Notes de Calcul
        "Note de Calcul d'Eclairement": 'badge-purple',
        'Note de Calcul BT': 'badge-purple',
        'Note de Calcul MT': 'badge-purple',
        'Note de Calcul DC': 'badge-purple',
        'Note de Calcul Batterie': 'badge-purple',
        'Note de Calcul CDC': 'badge-purple',
        'Note de Calcul de mise à la terre': 'badge-purple',
        'Note de Calcul paratonnerre': 'badge-purple',
        # Bilans
        'Bilan de puissance': 'badge-amber',
        'Bilan Thermique': 'badge-amber',
        # Réseaux & Synoptiques
        'Synoptiques CFO': 'badge-gray',
        'Synoptiques CFA': 'badge-gray',
        'Réseau sous dallage': 'badge-gray',
        'Réseau extérieur': 'badge-gray',
        'Réseau wifi': 'badge-gray',
        'Précâblage informatique': 'badge-gray',
        # Carnets
        'Carnet des câbles CFO': 'badge-gray',
        'Carnet des câbles CFA': 'badge-gray',
        # Autres
        'Prises de courant et forces': 'badge-blue',
        'Eclairage': 'badge-amber',
        'Réservations': 'badge-gray',
        'Mise à la terre et protection foudre': 'badge-gray',
        'Photovoltaïque': 'badge-amber',
        'Sonorisation': 'badge-purple',
        'Audiovisuel': 'badge-purple',
        'Système sécurité incendie': 'badge-gray',
        'Vidéosurveillance': 'badge-gray',
        "Contrôle d'accès et d'intrusion": 'badge-gray',
        "Gestion d'éclairage": 'badge-amber',
        'Aménagement local PCS': 'badge-blue',
        'Gestion domotique': 'badge-purple',
        'GTC': 'badge-purple',
        'TBE': 'badge-blue',
        'Congé': 'badge-amber',
    }

    @property
    def badge_class(self):
        return self.BADGE_MAP.get(self.type_activite, 'badge-gray')

    @property
    def projet_display(self):
        return self.projet.display_name if self.projet else 'Interne'
    
    @property
    def is_conge(self):
        """Vérifie si cette saisie est un congé"""
        return self.type_activite == 'Congé'
    
    @property
    def is_conge_multi_jours(self):
        """Vérifie si c'est un congé de plusieurs jours"""
        return self.is_conge and self.conge_date_debut is not None and self.conge_date_fin is not None

class Conge(db.Model):
    __tablename__ = 'conges'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    type_conge = db.Column(db.String(100), nullable=False)
    date_debut = db.Column(db.Date, nullable=False)
    date_fin = db.Column(db.Date, nullable=False)
    nb_jours = db.Column(db.Integer, nullable=False)
    motif = db.Column(db.Text)
    statut = db.Column(db.String(50), default='en_attente')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Notification(db.Model):
    __tablename__ = 'notifications'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    titre = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    couleur = db.Column(db.String(20), default='blue')
    lu = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    @property
    def time_ago(self):
        diff = datetime.utcnow() - self.created_at
        if diff.seconds < 3600:
            return f"Il y a {diff.seconds // 60}min"
        if diff.days == 0:
            return f"Il y a {diff.seconds // 3600}h"
        if diff.days == 1:
            return "Hier"
        return self.created_at.strftime('%d/%m/%Y')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ─────────────────────────────────────────
# AUTH ROUTES
# ─────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password, password):
            login_user(user, remember=True)
            return redirect(url_for('dashboard'))
        flash('Email ou mot de passe incorrect.', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ─────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    month_start = today.replace(day=1)

    heures_semaine = db.session.query(func.sum(Saisie.heures)).filter(
        Saisie.user_id == current_user.id,
        Saisie.date >= week_start,
        Saisie.date <= week_end
    ).scalar() or 0

    heures_mois = db.session.query(func.sum(Saisie.heures)).filter(
        Saisie.user_id == current_user.id,
        Saisie.date >= month_start,
        Saisie.date <= today
    ).scalar() or 0

    projets_actifs = Projet.query.filter_by(statut='actif').count()

    saisies_semaine = Saisie.query.filter(
        Saisie.user_id == current_user.id,
        Saisie.date >= week_start,
        Saisie.date <= week_end
    ).count()

    week_bars = []
    day_labels = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']
    for i in range(7):
        d = week_start + timedelta(days=i)
        h = db.session.query(func.sum(Saisie.heures)).filter(
            Saisie.user_id == current_user.id,
            Saisie.date == d
        ).scalar() or 0
        week_bars.append({
            'label': day_labels[i],
            'hours': float(h),
            'today': d == today,
            'weekend': i >= 5,
        })

    recent_saisies = Saisie.query.filter_by(user_id=current_user.id)\
       .order_by(Saisie.date.desc(), Saisie.created_at.desc()).limit(5).all()

    return render_template('dashboard.html',
        heures_semaine=heures_semaine,
        heures_mois=heures_mois,
        projets_actifs=projets_actifs,
        saisies_semaine=saisies_semaine,
        week_bars=week_bars,
        recent_saisies=recent_saisies,
        today=today,
    )

# ─────────────────────────────────────────
# SAISIES
# ─────────────────────────────────────────

@app.route('/saisie', methods=['GET', 'POST'])
@login_required
def saisie():
    """
    Route pour créer une nouvelle saisie (activité ou congé).
    Gère les deux modes de congé : 1 jour ou plusieurs jours.
    """
    projets = Projet.query.filter_by(statut='actif').order_by(Projet.nom).all()

    if request.method == 'POST':
        type_activite = request.form['type_activite']
        
        # ── Traitement spécifique pour les congés ──
        if type_activite == 'Congé':
            conge_duree_type = request.form.get('conge_duree_type', '1jour')
            
            if conge_duree_type == 'multi':
                # Mode "Plusieurs jours"
                date_debut_str = request.form['conge_date_debut']
                date_fin_str = request.form['conge_date_fin']
                date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
                date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
                # Bloquer les dates week-end
                if date_debut.weekday() in (5, 6) or date_fin.weekday() in (5, 6):
                    flash('Les dates de congé ne peuvent pas tomber un samedi ou un dimanche.', 'error')
                    return redirect(url_for('saisie'))
                nb_jours = (date_fin - date_debut).days + 1
                
                selected_date = date_debut
                heures = float(nb_jours)
                conge_date_debut = date_debut
                conge_date_fin = date_fin
            else:
                # Mode "≤ 1 jour"
                selected_date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
                # Bloquer les dates week-end
                if selected_date.weekday() in (5, 6):
                    flash('La date saisie ne peut pas tomber un samedi ou un dimanche.', 'error')
                    return redirect(url_for('saisie'))
                heures = float(request.form.get('conge_heures_1jour', 8))
                conge_date_debut = None
                conge_date_fin = None
            
            projet_id = None  # Pas de projet pour les congés
            
        else:
            # ── Traitement pour les activités normales ──
            selected_date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
            # Bloquer les dates week-end
            if selected_date.weekday() in (5, 6):
                flash('La date saisie ne peut pas tomber un samedi ou un dimanche.', 'error')
                return redirect(url_for('saisie'))
            heures = float(request.form['heures'])
            projet_id = request.form.get('projet_id') or None
            conge_date_debut = None
            conge_date_fin = None
        
        # ── Création de la saisie ──
        new_saisie = Saisie(
            user_id=current_user.id,
            projet_id=projet_id,
            date=selected_date,
            type_activite=type_activite,
            heures=heures,
            commentaire=request.form.get('commentaire', ''),
            statut='en_attente',
            conge_date_debut=conge_date_debut,
            conge_date_fin=conge_date_fin
        )
        db.session.add(new_saisie)
        
        # Notification
        if type_activite == 'Congé':
            desc = f'Congé de {heures} jour(s) ajouté.'
        else:
            desc = f'Saisie "{type_activite}" de {heures}h ajoutée.'
        
        notif = Notification(
            user_id=current_user.id,
            titre='Saisie enregistrée',
            description=desc,
            couleur='green',
        )
        db.session.add(notif)
        
        db.session.commit()
        flash('Saisie enregistrée avec succès!', 'success')
        return redirect(url_for('historique'))
    
    return render_template('saisie.html', projets=projets, today=date.today())


@app.route('/saisie/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def saisie_edit(id):
    """
    Route pour modifier une saisie existante.
    Gère les deux modes de congé : 1 jour ou plusieurs jours.
    """
    s = Saisie.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    projets = Projet.query.filter_by(statut='actif').order_by(Projet.nom).all()
    
    if request.method == 'POST':
        type_activite = request.form['type_activite']
        
        # ── Traitement spécifique pour les congés ──
        if type_activite == 'Congé':
            conge_duree_type = request.form.get('conge_duree_type', '1jour')
            
            if conge_duree_type == 'multi':
                # Mode "Plusieurs jours"
                date_debut_str = request.form['conge_date_debut']
                date_fin_str = request.form['conge_date_fin']
                date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
                date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
                nb_jours = (date_fin - date_debut).days + 1
                
                s.date = date_debut               # Date principale = date de début
                s.heures = float(nb_jours)         # 1 jour = 1 heure dans ce système
                s.conge_date_debut = date_debut    # Stocker explicitement
                s.conge_date_fin = date_fin        # Stocker explicitement
            else:
                # Mode "≤ 1 jour"
                s.date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
                s.heures = float(request.form.get('conge_heures_1jour', 8))
                s.conge_date_debut = None
                s.conge_date_fin = None
            
            s.projet_id = None  # Pas de projet pour les congés
            
        else:
            # ── Traitement pour les activités normales ──
            s.date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
            s.heures = float(request.form['heures'])
            s.projet_id = request.form.get('projet_id') or None
            s.conge_date_debut = None
            s.conge_date_fin = None
        
        s.type_activite = type_activite
        s.commentaire = request.form.get('commentaire', '')
        
        db.session.commit()
        flash('Saisie modifiée avec succès!', 'success')
        return redirect(url_for('historique'))
    
    return render_template('saisie.html', saisie=s, projets=projets, today=date.today())


@app.route('/saisie/<int:id>/delete', methods=['POST'])
@login_required
def saisie_delete(id):
    s = Saisie.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    db.session.delete(s)
    db.session.commit()
    flash('Saisie supprimée.', 'info')
    return redirect(url_for('historique'))


@app.route('/historique')
@login_required
def historique():
    filter_type = request.args.get('type', '')
    filter_projet = request.args.get('projet', '')
    filter_date_debut = request.args.get('date_debut', '')
    filter_date_fin = request.args.get('date_fin', '')
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    query = Saisie.query.filter_by(user_id=current_user.id)
    
    if filter_type:
        query = query.filter(Saisie.type_activite == filter_type)
    if filter_projet and filter_projet.isdigit():
        query = query.filter(Saisie.projet_id == int(filter_projet))
    if filter_date_debut:
        try:
            query = query.filter(Saisie.date >= datetime.strptime(filter_date_debut, '%Y-%m-%d').date())
        except ValueError:
            pass
    if filter_date_fin:
        try:
            query = query.filter(Saisie.date <= datetime.strptime(filter_date_fin, '%Y-%m-%d').date())
        except ValueError:
            pass
    
    all_saisies = query.order_by(Saisie.date.desc(), Saisie.created_at.desc()).all()
    
    total_saisies = len(all_saisies)
    total_heures = sum(float(s.heures) for s in all_saisies)
    
    total_pages = (total_saisies + per_page - 1) // per_page if total_saisies > 0 else 1
    start = (page - 1) * per_page
    end = start + per_page
    page_saisies = all_saisies[start:end]
    page_total_heures = sum(float(s.heures) for s in page_saisies)
    
    start_index = start + 1 if total_saisies > 0 else 0
    end_index = min(end, total_saisies)
    
    projets = Projet.query.filter_by(statut='actif').order_by(Projet.nom).all()
    
    notifs = Notification.query.filter_by(user_id=current_user.id)\
        .order_by(Notification.created_at.desc()).limit(10).all()
    notif_unread = sum(1 for n in notifs if not n.lu)
    
    return render_template('historique.html',
                         saisies=all_saisies,
                         page_saisies=page_saisies,
                         page=page,
                         total_pages=total_pages,
                         total_saisies=total_saisies,
                         total_heures=total_heures,
                         page_total_heures=page_total_heures,
                         start_index=start_index,
                         end_index=end_index,
                         projets=projets,
                         filter_type=filter_type,
                         filter_projet=filter_projet,
                         filter_date_debut=filter_date_debut,
                         filter_date_fin=filter_date_fin,
                         notifs=notifs,
                         notif_unread=notif_unread)

# ─────────────────────────────────────────
# FILTER HELPER FUNCTIONS
# ─────────────────────────────────────────

def get_filtered_saisies():
    """Fonction utilitaire pour récupérer les saisies filtrées"""
    filter_type = request.args.get('type', '')
    filter_projet = request.args.get('projet', '')
    filter_date_debut = request.args.get('date_debut', '')
    filter_date_fin = request.args.get('date_fin', '')
    
    query = Saisie.query.filter_by(user_id=current_user.id)
    
    if filter_type:
        query = query.filter(Saisie.type_activite == filter_type)
    if filter_projet and filter_projet.isdigit():
        query = query.filter(Saisie.projet_id == int(filter_projet))
    if filter_date_debut:
        try:
            date_debut = datetime.strptime(filter_date_debut, '%Y-%m-%d').date()
            query = query.filter(Saisie.date >= date_debut)
        except ValueError:
            pass
    if filter_date_fin:
        try:
            date_fin = datetime.strptime(filter_date_fin, '%Y-%m-%d').date()
            query = query.filter(Saisie.date <= date_fin)
        except ValueError:
            pass
    
    return query.order_by(Saisie.date.desc(), Saisie.created_at.desc()).all()

def get_filtered_conges():
    """Fonction utilitaire pour récupérer les congés filtrés"""
    filter_type = request.args.get('type', '')
    filter_statut = request.args.get('statut', '')
    filter_date_debut = request.args.get('date_debut', '')
    filter_date_fin = request.args.get('date_fin', '')
    
    query = Conge.query.filter_by(user_id=current_user.id)
    
    if filter_type:
        query = query.filter(Conge.type_conge == filter_type)
    if filter_statut:
        query = query.filter(Conge.statut == filter_statut)
    if filter_date_debut:
        try:
            date_debut = datetime.strptime(filter_date_debut, '%Y-%m-%d').date()
            query = query.filter(Conge.date_debut >= date_debut)
        except ValueError:
            pass
    if filter_date_fin:
        try:
            date_fin = datetime.strptime(filter_date_fin, '%Y-%m-%d').date()
            query = query.filter(Conge.date_fin <= date_fin)
        except ValueError:
            pass
    
    return query.order_by(Conge.created_at.desc()).all()


# ============================================
# API JSON pour Saisies et Congés
# ============================================

@app.route('/api/saisies')
@login_required
def api_saisies():
    """API JSON pour les saisies filtrées"""
    try:
        filter_type = request.args.get('type', '')
        filter_projet = request.args.get('projet', '')
        filter_date_debut = request.args.get('date_debut', '')
        filter_date_fin = request.args.get('date_fin', '')
        
        query = Saisie.query.filter_by(user_id=current_user.id)
        
        if filter_type:
            query = query.filter(Saisie.type_activite == filter_type)
        if filter_projet and filter_projet.isdigit():
            query = query.filter(Saisie.projet_id == int(filter_projet))
        if filter_date_debut:
            try:
                date_debut = datetime.strptime(filter_date_debut, '%Y-%m-%d').date()
                query = query.filter(Saisie.date >= date_debut)
            except ValueError:
                pass
        if filter_date_fin:
            try:
                date_fin = datetime.strptime(filter_date_fin, '%Y-%m-%d').date()
                query = query.filter(Saisie.date <= date_fin)
            except ValueError:
                pass
        
        saisies = query.order_by(Saisie.date.desc(), Saisie.created_at.desc()).all()
        
        return jsonify([{
            'id': s.id,
            'date': s.date.strftime('%d/%m/%Y'),
            'type': s.type_activite,
            'projet': s.projet_display,
            'heures': float(s.heures),
            'commentaire': s.commentaire,
            'statut': s.statut or 'en_attente',
            'conge_date_debut': s.conge_date_debut.isoformat() if s.conge_date_debut else None,
            'conge_date_fin': s.conge_date_fin.isoformat() if s.conge_date_fin else None,
        } for s in saisies])
        
    except Exception as e:
        print(f"Erreur API saisies: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/conges')
@login_required
def api_conges():
    """API JSON pour les congés filtrés"""
    try:
        filter_type = request.args.get('type', '')
        filter_statut = request.args.get('statut', '')
        filter_date_debut = request.args.get('date_debut', '')
        filter_date_fin = request.args.get('date_fin', '')
        
        query = Conge.query.filter_by(user_id=current_user.id)
        
        if filter_type:
            query = query.filter(Conge.type_conge == filter_type)
        if filter_statut:
            query = query.filter(Conge.statut == filter_statut)
        if filter_date_debut:
            try:
                date_debut = datetime.strptime(filter_date_debut, '%Y-%m-%d').date()
                query = query.filter(Conge.date_debut >= date_debut)
            except ValueError:
                pass
        if filter_date_fin:
            try:
                date_fin = datetime.strptime(filter_date_fin, '%Y-%m-%d').date()
                query = query.filter(Conge.date_fin <= date_fin)
            except ValueError:
                pass
        
        conges = query.order_by(Conge.created_at.desc()).all()
        
        return jsonify([{
            'id': c.id,
            'created_at': c.created_at.strftime('%d/%m/%Y'),
            'type': c.type_conge,
            'date_debut': c.date_debut.strftime('%d/%m/%Y'),
            'date_fin': c.date_fin.strftime('%d/%m/%Y'),
            'nb_jours': c.nb_jours,
            'motif': c.motif,
            'statut': c.statut
        } for c in conges])
        
    except Exception as e:
        print(f"Erreur API congés: {e}")
        return jsonify({'error': str(e)}), 500

# ─────────────────────────────────────────
# PDF HELPER
# ─────────────────────────────────────────

# Couleurs Futura
PDF_ACCENT   = colors.HexColor('#2563EB')
PDF_ACCENT2  = colors.HexColor('#1D4ED8')
PDF_HEADER   = colors.HexColor('#1E3A5F')
PDF_ROW_ALT  = colors.HexColor('#F0F5FF')
PDF_ROW_ODD  = colors.HexColor('#FFFFFF')
PDF_TOTAL    = colors.HexColor('#E8EFFE')
PDF_BORDER   = colors.HexColor('#CBD5E1')
PDF_TEXT     = colors.HexColor('#1A1A1A')
PDF_GRAY     = colors.HexColor('#64748B')

COMPANY_NAME = "Futura Expertise"

class PDFHeaderFooter:
    """Canvas mixin qui dessine header + footer sur chaque page."""
    def __init__(self, title, username, subtitle=""):
        self.title    = title
        self.username = username
        self.subtitle = subtitle

    def __call__(self, canv, doc):
        canv.saveState()
        W, H = landscape(A4)

        # ── Header band ──
        canv.setFillColor(PDF_ACCENT)
        canv.rect(0, H - 1.6*cm, W, 1.6*cm, fill=1, stroke=0)

        # Nom de société (gauche)
        canv.setFillColor(colors.white)
        canv.setFont('Helvetica-Bold', 11)
        canv.drawString(1*cm, H - 1.1*cm, COMPANY_NAME)

        # Titre document (centre)
        canv.setFont('Helvetica-Bold', 10)
        canv.drawCentredString(W/2, H - 1.1*cm, self.title)

        # Utilisateur (droite)
        canv.setFont('Helvetica', 9)
        canv.drawRightString(W - 1*cm, H - 1.1*cm, self.username)

        # Sous-ligne date (si subtitle)
        if self.subtitle:
            canv.setFont('Helvetica', 8)
            canv.setFillColor(colors.HexColor('#BFDBFE'))
            canv.drawCentredString(W/2, H - 1.45*cm, self.subtitle)

        # ── Footer ──
        canv.setFillColor(PDF_GRAY)
        canv.setFont('Helvetica', 8)
        canv.drawString(1*cm, 0.5*cm, f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
        canv.drawCentredString(W/2, 0.5*cm, COMPANY_NAME)
        canv.drawRightString(W - 1*cm, 0.5*cm, f"Page {doc.page}")

        # Ligne footer
        canv.setStrokeColor(PDF_BORDER)
        canv.setLineWidth(0.5)
        canv.line(1*cm, 0.8*cm, W - 1*cm, 0.8*cm)

        canv.restoreState()


def make_table_style(num_rows, has_total=True):
    """Retourne un TableStyle propre et moderne."""
    style = [
        # Header
        ('BACKGROUND',   (0, 0), (-1, 0),  PDF_HEADER),
        ('TEXTCOLOR',    (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0),  9),
        ('BOTTOMPADDING',(0, 0), (-1, 0),  8),
        ('TOPPADDING',   (0, 0), (-1, 0),  8),
        ('ALIGN',        (0, 0), (-1, 0),  'CENTER'),
        # Corps
        ('FONTNAME',     (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',     (0, 1), (-1, -1), 8),
        ('TOPPADDING',   (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING',(0, 1), (-1, -1), 5),
        ('ALIGN',        (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
        # Bordures
        ('LINEBELOW',    (0, 0), (-1, 0),  1, PDF_ACCENT2),
        ('INNERGRID',    (0, 1), (-1, -1), 0.4, PDF_BORDER),
        ('BOX',          (0, 0), (-1, -1), 0.8, PDF_ACCENT),
    ]
    # Lignes alternées
    for i in range(1, num_rows):
        if has_total and i == num_rows - 1:
            style += [
                ('BACKGROUND', (0, i), (-1, i), PDF_TOTAL),
                ('FONTNAME',   (0, i), (-1, i), 'Helvetica-Bold'),
                ('LINEABOVE',  (0, i), (-1, i), 0.8, PDF_ACCENT),
            ]
        elif i % 2 == 0:
            style.append(('BACKGROUND', (0, i), (-1, i), PDF_ROW_ALT))
        else:
            style.append(('BACKGROUND', (0, i), (-1, i), PDF_ROW_ODD))
    return TableStyle(style)


def pdf_section_title(text, styles):
    style = ParagraphStyle('SecTitle',
        fontName='Helvetica-Bold', fontSize=11,
        textColor=PDF_ACCENT, spaceBefore=16, spaceAfter=6)
    return Paragraph(text, style)


# ─────────────────────────────────────────
# EXPORT SAISIES (CSV, XLSX, PDF)
# ─────────────────────────────────────────

@app.route('/exporter/csv')
@login_required
def exporter_csv():
    try:
        saisies = get_filtered_saisies()
        
        output = StringIO()
        output.write('\uFEFF')
        writer = csv.writer(output, delimiter=';')
        
        writer.writerow(['Date', 'Type', 'Projet', 'Heures', 'Commentaire', 'Statut'])
        
        statut_map = {
            'en_attente': 'En attente',
            'valide': 'Validé',
            'approuve': 'Approuvé',
            'rejete': 'Rejeté',
            'refuse': 'Refusé'
        }
        
        for s in saisies:
            writer.writerow([
                s.date.strftime('%d/%m/%Y'),
                s.type_activite,
                s.projet_display,
                f"{s.heures}",
                s.commentaire or '',
                statut_map.get(s.statut, 'En attente')
            ])
        
        if saisies:
            total_heures = sum(float(s.heures) for s in saisies)
            writer.writerow([])
            writer.writerow(['Total général', '', '', f"{total_heures}h", '', ''])
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=saisies_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        return response
        
    except Exception as e:
        print(f"Erreur export CSV: {str(e)}")
        flash(f"Erreur lors de l'export: {str(e)}", 'error')
        return redirect(url_for('historique'))

@app.route('/exporter/xlsx')
@login_required
def exporter_xlsx():
    try:
        saisies = get_filtered_saisies()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Saisies"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        headers = ['Date', 'Type', 'Projet', 'Heures', 'Commentaire', 'Statut']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        statut_map = {
            'en_attente': 'En attente',
            'valide': 'Validé',
            'approuve': 'Approuvé',
            'rejete': 'Rejeté',
            'refuse': 'Refusé'
        }
        
        for row, s in enumerate(saisies, 2):
            ws.cell(row=row, column=1, value=s.date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=s.type_activite)
            ws.cell(row=row, column=3, value=s.projet_display)
            ws.cell(row=row, column=4, value=float(s.heures))
            ws.cell(row=row, column=5, value=s.commentaire or '')
            ws.cell(row=row, column=6, value=statut_map.get(s.statut, 'En attente'))
        
        if len(saisies) > 0:
            total_row = len(saisies) + 2
            total_heures = sum(float(s.heures) for s in saisies)
            ws.cell(row=total_row, column=3, value="TOTAL")
            ws.cell(row=total_row, column=4, value=float(total_heures))
            for col in [3, 4]:
                cell = ws.cell(row=total_row, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="right", vertical="center")
        
        for col in range(1, 7):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        for row in range(2, len(saisies) + 3):
            cell = ws.cell(row=row, column=4)
            cell.alignment = Alignment(horizontal="right", vertical="center")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=saisies_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
        
    except Exception as e:
        print(f"Erreur export XLSX: {str(e)}")
        flash(f"Erreur lors de l'export: {str(e)}", 'error')
        return redirect(url_for('historique'))

@app.route('/exporter/pdf')
@login_required
def exporter_pdf():
    try:
        saisies = get_filtered_saisies()
        total_heures = sum(float(s.heures) for s in saisies)

        buffer = BytesIO()
        username = f"{current_user.prenom} {current_user.nom}"
        hf = PDFHeaderFooter(
            title="Historique des saisies",
            username=username,
            subtitle=f"Période : {datetime.now().strftime('%B %Y')}  •  {len(saisies)} enregistrement(s)"
        )
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                rightMargin=1.2*cm, leftMargin=1.2*cm,
                                topMargin=2.2*cm, bottomMargin=1.4*cm)

        statut_map = {
            'en_attente': 'En attente', 'valide': 'Validé',
            'approuve': 'Approuvé', 'rejete': 'Rejeté', 'refuse': 'Refusé'
        }

        data = [['Date', "Type d'activité", 'Projet', 'Heures', 'Commentaire', 'Statut']]
        for s in saisies:
            data.append([
                s.date.strftime('%d/%m/%Y'),
                s.type_activite,
                s.projet_display[:35],
                f"{s.heures}h",
                (s.commentaire or '—')[:45],
                statut_map.get(s.statut, 'En attente')
            ])
        data.append(['', '', 'TOTAL', f"{total_heures}h", '', ''])

        col_w = [2.2*cm, 6*cm, 5*cm, 1.8*cm, 8*cm, 2.8*cm]
        table = Table(data, colWidths=col_w, repeatRows=1)
        table.setStyle(make_table_style(len(data)))

        doc.build([table], onFirstPage=hf, onLaterPages=hf)
        pdf = buffer.getvalue()
        buffer.close()

        response = make_response(pdf)
        response.headers['Content-Disposition'] = f'attachment; filename=saisies_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        return response

    except Exception as e:
        print(f"Erreur export PDF: {str(e)}")
        flash(f"Erreur lors de l'export: {str(e)}", 'error')
        return redirect(url_for('historique'))

# ─────────────────────────────────────────
# EXPORT CONGÉS (CSV, XLSX, PDF)
# ─────────────────────────────────────────

@app.route('/exporter/conges/csv')
@login_required
def exporter_conges_csv():
    try:
        conges = get_filtered_conges()
        
        output = StringIO()
        output.write('\uFEFF')
        writer = csv.writer(output, delimiter=';')
        
        writer.writerow(['Date demande', 'Type', 'Date début', 'Date fin', 'Nb jours', 'Motif', 'Statut'])
        
        statut_map = {
            'en_attente': 'En attente',
            'approuve': 'Approuvé',
            'refuse': 'Refusé'
        }
        
        for c in conges:
            writer.writerow([
                c.created_at.strftime('%d/%m/%Y'),
                c.type_conge,
                c.date_debut.strftime('%d/%m/%Y'),
                c.date_fin.strftime('%d/%m/%Y'),
                c.nb_jours,
                c.motif or '',
                statut_map.get(c.statut, 'En attente')
            ])
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=conges_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        return response
        
    except Exception as e:
        print(f"Erreur export congés CSV: {e}")
        flash("Erreur lors de l'export des congés", 'error')
        return redirect(url_for('rapports'))

@app.route('/exporter/conges/xlsx')
@login_required
def exporter_conges_xlsx():
    try:
        conges = get_filtered_conges()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Congés"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        headers = ['Date demande', 'Type', 'Date début', 'Date fin', 'Nb jours', 'Motif', 'Statut']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        statut_map = {
            'en_attente': 'En attente',
            'approuve': 'Approuvé',
            'refuse': 'Refusé'
        }
        
        for row, c in enumerate(conges, 2):
            ws.cell(row=row, column=1, value=c.created_at.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=c.type_conge)
            ws.cell(row=row, column=3, value=c.date_debut.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=4, value=c.date_fin.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=5, value=c.nb_jours)
            ws.cell(row=row, column=6, value=c.motif or '')
            ws.cell(row=row, column=7, value=statut_map.get(c.statut, 'En attente'))
        
        for col in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=conges_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
        
    except Exception as e:
        print(f"Erreur export congés XLSX: {e}")
        flash("Erreur lors de l'export des congés", 'error')
        return redirect(url_for('rapports'))

@app.route('/exporter/conges/pdf')
@login_required
def exporter_conges_pdf():
    try:
        conges = get_filtered_conges()

        buffer = BytesIO()
        username = f"{current_user.prenom} {current_user.nom}"
        hf = PDFHeaderFooter(
            title="Demandes de congé",
            username=username,
            subtitle=f"{len(conges)} demande(s)  •  {datetime.now().strftime('%B %Y')}"
        )
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                rightMargin=1.2*cm, leftMargin=1.2*cm,
                                topMargin=2.2*cm, bottomMargin=1.4*cm)

        statut_map = {
            'en_attente': 'En attente',
            'approuve': 'Approuvé',
            'refuse': 'Refusé'
        }

        data = [['Date demande', 'Type congé', 'Date début', 'Date fin', 'Nb jours', 'Motif', 'Statut']]
        for c in conges:
            data.append([
                c.created_at.strftime('%d/%m/%Y'),
                c.type_conge,
                c.date_debut.strftime('%d/%m/%Y'),
                c.date_fin.strftime('%d/%m/%Y'),
                str(c.nb_jours),
                (c.motif or '—')[:40],
                statut_map.get(c.statut, 'En attente')
            ])

        col_w = [2.5*cm, 4*cm, 2.5*cm, 2.5*cm, 2*cm, 8*cm, 2.5*cm]
        table = Table(data, colWidths=col_w, repeatRows=1)
        table.setStyle(make_table_style(len(data), has_total=False))

        doc.build([table], onFirstPage=hf, onLaterPages=hf)
        pdf = buffer.getvalue()
        buffer.close()

        response = make_response(pdf)
        response.headers['Content-Disposition'] = f'attachment; filename=conges_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        return response

    except Exception as e:
        print(f"Erreur export congés PDF: {e}")
        flash("Erreur lors de l'export des congés", 'error')
        return redirect(url_for('rapports'))

# ─────────────────────────────────────────
# RAPPORTS
# ─────────────────────────────────────────

@app.route('/rapports')
@login_required
def rapports():
    projets = Projet.query.filter_by(statut='actif').order_by(Projet.nom).all()
    
    notifs = Notification.query.filter_by(user_id=current_user.id)\
        .order_by(Notification.created_at.desc()).limit(10).all()
    notif_unread = sum(1 for n in notifs if not n.lu)
    
    return render_template('rapports.html',
                         projets=projets,
                         notifs=notifs,
                         notif_unread=notif_unread)

# ─────────────────────────────────────────
# CONGÉS
# ─────────────────────────────────────────

@app.route('/conge', methods=['GET', 'POST'])
@login_required
def conge():
    if request.method == 'POST':
        date_debut = datetime.strptime(request.form['date_debut'], '%Y-%m-%d').date()
        date_fin   = datetime.strptime(request.form['date_fin'], '%Y-%m-%d').date()
        # Bloquer les dates week-end (samedi=5, dimanche=6)
        if date_debut.weekday() in (5, 6) or date_fin.weekday() in (5, 6):
            flash('Les dates de congé ne peuvent pas tomber un samedi ou un dimanche.', 'error')
            return redirect(url_for('conge'))
        new_conge = Conge(
            user_id = current_user.id,
            type_conge = request.form['type_conge'],
            date_debut = date_debut,
            date_fin = date_fin,
            nb_jours = int(request.form['nb_jours']),
            motif = request.form.get('motif', ''),
        )
        db.session.add(new_conge)
        notif = Notification(
            user_id = current_user.id,
            titre = 'Demande de congé envoyée',
            description = f'Demande de {new_conge.nb_jours} jour(s) soumise, en attente d\'approbation.',
            couleur = 'amber',
        )
        db.session.add(notif)
        db.session.commit()
        flash('Demande de congé envoyée avec succès!', 'success')
        return redirect(url_for('dashboard'))

    mes_conges = Conge.query.filter_by(user_id=current_user.id)\
       .order_by(Conge.created_at.desc()).all()
    return render_template('conge.html', mes_conges=mes_conges)

# ─────────────────────────────────────────
# API JSON
# ─────────────────────────────────────────

@app.route('/api/notifications')
@login_required
def api_notifications():
    notifs = Notification.query.filter_by(user_id=current_user.id)\
       .order_by(Notification.created_at.desc()).limit(15).all()
    return jsonify([{
        'id': n.id,
        'titre': n.titre,
        'description': n.description,
        'couleur': n.couleur,
        'lu': n.lu,
        'time': n.time_ago,
    } for n in notifs])



@app.route('/api/notifications/<int:id>/read', methods=['POST'])
@login_required
def api_notif_read(id):
    n = Notification.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    n.lu = True
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def api_notif_read_all():
    Notification.query.filter_by(user_id=current_user.id, lu=False)\
       .update({'lu': True})
    db.session.commit()
    return jsonify({'ok': True})

# ─────────────────────────────────────────
# RAPPORT COMPLET (SAISIES + CONGÉS)
# ─────────────────────────────────────────

@app.route('/exporter/rapport-complet/csv')
@login_required
def exporter_rapport_complet_csv():
    try:
        saisies = get_filtered_saisies()
        conges = get_filtered_conges()
        
        output = StringIO()
        output.write('\uFEFF')
        writer = csv.writer(output, delimiter=';')
        
        # Section Saisies
        writer.writerow(['=== SAISIES ==='])
        writer.writerow(['Date', 'Type', 'Projet', 'Heures', 'Commentaire', 'Statut'])
        
        statut_map = {
            'en_attente': 'En attente',
            'valide': 'Validé',
            'approuve': 'Approuvé',
            'rejete': 'Rejeté',
            'refuse': 'Refusé'
        }
        
        for s in saisies:
            writer.writerow([
                s.date.strftime('%d/%m/%Y'),
                s.type_activite,
                s.projet_display,
                f"{s.heures}",
                s.commentaire or '',
                statut_map.get(s.statut, 'En attente')
            ])
        
        total_heures = sum(float(s.heures) for s in saisies)
        writer.writerow(['Total heures', '', '', f"{total_heures}h", '', ''])
        writer.writerow([])
        
        # Section Congés
        writer.writerow(['=== CONGÉS ==='])
        writer.writerow(['Date demande', 'Type', 'Date début', 'Date fin', 'Nb jours', 'Motif', 'Statut'])
        
        statut_conge_map = {
            'en_attente': 'En attente',
            'approuve': 'Approuvé',
            'refuse': 'Refusé'
        }
        
        for c in conges:
            writer.writerow([
                c.created_at.strftime('%d/%m/%Y'),
                c.type_conge,
                c.date_debut.strftime('%d/%m/%Y'),
                c.date_fin.strftime('%d/%m/%Y'),
                c.nb_jours,
                c.motif or '',
                statut_conge_map.get(c.statut, 'En attente')
            ])
        
        total_jours = sum(c.nb_jours for c in conges)
        writer.writerow(['Total jours', '', '', '', f"{total_jours} jours", '', ''])
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=rapport_complet_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        return response
        
    except Exception as e:
        print(f"Erreur export rapport complet CSV: {e}")
        flash("Erreur lors de l'export du rapport complet", 'error')
        return redirect(url_for('rapports'))

@app.route('/exporter/rapport-complet/pdf')
@login_required
def exporter_rapport_complet_pdf():
    try:
        saisies = get_filtered_saisies()
        conges  = get_filtered_conges()
        total_heures = sum(float(s.heures) for s in saisies)
        total_jours  = sum(c.nb_jours for c in conges)

        buffer = BytesIO()
        username = f"{current_user.prenom} {current_user.nom}"
        hf = PDFHeaderFooter(
            title="Rapport complet",
            username=username,
            subtitle=f"{len(saisies)} saisie(s)  •  {len(conges)} congé(s)  •  {datetime.now().strftime('%B %Y')}"
        )
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                rightMargin=1.2*cm, leftMargin=1.2*cm,
                                topMargin=2.2*cm, bottomMargin=1.4*cm)

        statut_map = {
            'en_attente': 'En attente', 'valide': 'Validé',
            'approuve': 'Approuvé', 'rejete': 'Rejeté', 'refuse': 'Refusé'
        }
        statut_conge_map = {
            'en_attente': 'En attente', 'approuve': 'Approuvé', 'refuse': 'Refusé'
        }

        elements = []

        # ── Section 1 : Saisies ──
        elements.append(pdf_section_title("1.  Historique des saisies", getSampleStyleSheet()))
        elements.append(HRFlowable(width="100%", thickness=1, color=PDF_ACCENT, spaceAfter=8))

        data_s = [['Date', "Type d'activité", 'Projet', 'Heures', 'Commentaire', 'Statut']]
        for s in saisies:
            data_s.append([
                s.date.strftime('%d/%m/%Y'),
                s.type_activite,
                s.projet_display[:30],
                f"{s.heures}h",
                (s.commentaire or '—')[:40],
                statut_map.get(s.statut, 'En attente')
            ])
        data_s.append(['', '', 'TOTAL', f"{total_heures}h", '', ''])

        col_s = [2.2*cm, 5.5*cm, 4.5*cm, 1.8*cm, 7.5*cm, 2.5*cm]
        t_s = Table(data_s, colWidths=col_s, repeatRows=1)
        t_s.setStyle(make_table_style(len(data_s)))
        elements.append(t_s)
        elements.append(Spacer(1, 0.6*cm))

        # ── Section 2 : Congés ──
        elements.append(pdf_section_title("2.  Demandes de congé", getSampleStyleSheet()))
        elements.append(HRFlowable(width="100%", thickness=1, color=PDF_ACCENT, spaceAfter=8))

        data_c = [['Date demande', 'Type congé', 'Date début', 'Date fin', 'Nb jours', 'Motif', 'Statut']]
        for c in conges:
            data_c.append([
                c.created_at.strftime('%d/%m/%Y'),
                c.type_conge,
                c.date_debut.strftime('%d/%m/%Y'),
                c.date_fin.strftime('%d/%m/%Y'),
                str(c.nb_jours),
                (c.motif or '—')[:35],
                statut_conge_map.get(c.statut, 'En attente')
            ])
        data_c.append(['', '', '', 'TOTAL', f"{total_jours} j", '', ''])

        col_c = [2.5*cm, 3.5*cm, 2.5*cm, 2.5*cm, 2*cm, 8*cm, 2.5*cm]
        t_c = Table(data_c, colWidths=col_c, repeatRows=1)
        t_c.setStyle(make_table_style(len(data_c)))
        elements.append(t_c)

        doc.build(elements, onFirstPage=hf, onLaterPages=hf)
        pdf = buffer.getvalue()
        buffer.close()

        response = make_response(pdf)
        response.headers['Content-Disposition'] = f'attachment; filename=rapport_complet_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response.headers['Content-Type'] = 'application/pdf'
        return response

    except Exception as e:
        print(f"Erreur export rapport complet PDF: {e}")
        flash("Erreur lors de l'export du rapport complet", 'error')
        return redirect(url_for('rapports'))

@app.route('/exporter/rapport-complet/xlsx')
@login_required
def exporter_rapport_complet_xlsx():
    try:
        saisies = get_filtered_saisies()
        conges = get_filtered_conges()
        
        wb = openpyxl.Workbook()
        
        # ========== SHEET 1: SAISIES ==========
        ws1 = wb.active
        ws1.title = "Saisies"
        
        ws1['A1'] = 'Date'
        ws1['B1'] = 'Type'
        ws1['C1'] = 'Projet'
        ws1['D1'] = 'Heures'
        ws1['E1'] = 'Commentaire'
        ws1['F1'] = 'Statut'
        
        statut_map = {
            'en_attente': 'En attente',
            'valide': 'Validé',
            'approuve': 'Approuvé',
            'rejete': 'Rejeté',
            'refuse': 'Refusé'
        }
        
        row = 2
        for s in saisies:
            ws1[f'A{row}'] = s.date.strftime('%d/%m/%Y')
            ws1[f'B{row}'] = s.type_activite
            ws1[f'C{row}'] = s.projet_display
            ws1[f'D{row}'] = float(s.heures)
            ws1[f'E{row}'] = s.commentaire or ''
            ws1[f'F{row}'] = statut_map.get(s.statut, 'En attente')
            row += 1
        
        if saisies:
            total_heures = sum(float(s.heures) for s in saisies)
            ws1[f'C{row}'] = 'TOTAL'
            ws1[f'D{row}'] = total_heures
        
        # ========== SHEET 2: CONGÉS ==========
        ws2 = wb.create_sheet(title="Conges")
        
        ws2['A1'] = 'Date demande'
        ws2['B1'] = 'Type'
        ws2['C1'] = 'Date début'
        ws2['D1'] = 'Date fin'
        ws2['E1'] = 'Nb jours'
        ws2['F1'] = 'Motif'
        ws2['G1'] = 'Statut'
        
        statut_conge_map = {
            'en_attente': 'En attente',
            'approuve': 'Approuvé',
            'refuse': 'Refusé'
        }
        
        row = 2
        for c in conges:
            ws2[f'A{row}'] = c.created_at.strftime('%d/%m/%Y')
            ws2[f'B{row}'] = c.type_conge
            ws2[f'C{row}'] = c.date_debut.strftime('%d/%m/%Y')
            ws2[f'D{row}'] = c.date_fin.strftime('%d/%m/%Y')
            ws2[f'E{row}'] = c.nb_jours
            ws2[f'F{row}'] = c.motif or ''
            ws2[f'G{row}'] = statut_conge_map.get(c.statut, 'En attente')
            row += 1
        
        if conges:
            total_jours = sum(c.nb_jours for c in conges)
            ws2[f'E{row}'] = 'TOTAL'
            ws2[f'F{row}'] = f"{total_jours} jours"
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=rapport_complet_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
        
    except Exception as e:
        print(f"Erreur export rapport complet XLSX: {e}")
        flash(f"Erreur: {str(e)}", 'error')
        return redirect(url_for('rapports'))

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, port=5000)   